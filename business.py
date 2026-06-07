from dataclasses import dataclass
from datetime import date, datetime, timedelta

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from constants import (
    DAILY_WFH_LIMIT,
    DATA_DIR,
    EMPLOYEES,
    METADATA_HEADERS,
    SHEET_HEADERS,
    VISIBLE_HEADERS,
    WORKBOOK_TITLE,
)


@dataclass
class RequestResult:
    message: str
    category: str = "success"


NAME_COLUMN = VISIBLE_HEADERS.index("Name") + 1
TID_COLUMN = VISIBLE_HEADERS.index("TID") + 1
WEEKDAY_START_COLUMN = VISIBLE_HEADERS.index("Mon") + 1
WEEKDAY_LABELS = ["Mon", "Tue", "Wed", "Thu", "Fri"]
WEEKDAY_COLUMNS = {
    index: WEEKDAY_START_COLUMN + index
    for index in range(len(WEEKDAY_LABELS))
}
METADATA_START_COLUMN = len(VISIBLE_HEADERS) + 1
EMAIL_COLUMN = METADATA_START_COLUMN + METADATA_HEADERS.index("Mail ID")
WEEK_START_COLUMN = METADATA_START_COLUMN + METADATA_HEADERS.index("Week Start")
WEEK_END_COLUMN = METADATA_START_COLUMN + METADATA_HEADERS.index("Week End")
STATUS_COLUMN = METADATA_START_COLUMN + METADATA_HEADERS.index("Status")
RANGE_FILE_PATTERN = "*_to_*.xlsx"
USERS_FILE = DATA_DIR / "users.xlsx"
USER_HEADERS = ["Name", "TID", "Mail ID"]


def normalize_text(value):
    return str(value or "").strip()


def create_users_workbook():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Users"

    for column, header in enumerate(USER_HEADERS, start=1):
        cell = sheet.cell(row=1, column=column, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for employee in EMPLOYEES:
        sheet.append([
            normalize_text(employee.get("name")),
            normalize_text(employee.get("tid")),
            normalize_text(employee.get("email")),
        ])

    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 14
    sheet.column_dimensions["C"].width = 30
    workbook.save(USERS_FILE)
    return workbook


def get_users_workbook():
    if USERS_FILE.exists():
        return load_workbook(USERS_FILE)
    return create_users_workbook()


def get_employees():
    workbook = get_users_workbook()
    sheet = workbook.active
    employees = []

    for row in sheet.iter_rows(min_row=2, max_col=len(USER_HEADERS), values_only=True):
        name, tid, email = [normalize_text(value) for value in row]
        if name:
            employees.append({"name": name, "tid": tid, "email": email})

    return employees


def get_user_records():
    return [
        {**employee, "index": index}
        for index, employee in enumerate(get_employees())
    ]


def get_user_sheet_row(sheet, user_index):
    current_index = 0

    for row_number in range(2, sheet.max_row + 1):
        name = normalize_text(sheet.cell(row=row_number, column=1).value)
        if not name:
            continue

        if current_index == user_index:
            return row_number

        current_index += 1

    return None


def get_range_file(week_start, week_end):
    return DATA_DIR / f"{week_start.isoformat()}_to_{week_end.isoformat()}.xlsx"


def get_latest_data_file():
    files = [
        file_path
        for file_path in DATA_DIR.glob(RANGE_FILE_PATTERN)
        if file_path.is_file()
    ]

    if not files:
        return None

    return max(files, key=lambda file_path: file_path.stat().st_mtime)


def current_week_range():
    today = date.today()
    start = today - timedelta(days=today.weekday())
    end = start + timedelta(days=6)
    return start, end


def next_week_range():
    current_week_start, _ = current_week_range()
    start = current_week_start + timedelta(days=7)
    end = start + timedelta(days=6)
    return start, end


def parse_date(date_text):
    try:
        return datetime.strptime(date_text, "%Y-%m-%d").date()
    except (TypeError, ValueError):
        return None


def parse_week_range(week_start_text, week_end_text):
    week_start = parse_date(week_start_text)
    week_end = parse_date(week_end_text)

    if not week_start or not week_end:
        return None, None

    return week_start, week_end


def validate_week_range(week_start, week_end):
    if not week_start or not week_end:
        return "Please select a valid start date and end date."

    if week_end < week_start:
        return "End date must be after start date."

    if (week_end - week_start).days != 6:
        return "Please select a 7-day date range, for example 2026-06-08 to 2026-06-14."

    return None


def get_week_label(week_start):
    return f"WK{week_start.isocalendar().week}"


def get_working_week_dates(week_start=None):
    start = week_start or next_week_range()[0]
    return [start + timedelta(days=offset) for offset in range(5)]


def get_weekday_headers(week_start=None):
    return [
        f"{day.strftime('%a')} {day.strftime('%d-%b')}"
        for day in get_working_week_dates(week_start)
    ]


def style_worksheet(sheet, week_start=None):
    title_fill = PatternFill("solid", fgColor="0070C0")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    metadata_fill = PatternFill("solid", fgColor="F2F4F7")
    thin_side = Side(style="thin", color="808080")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    visible_end_column = get_column_letter(len(VISIBLE_HEADERS))
    active_week_start = week_start or next_week_range()[0]

    sheet.title = "WFH Plan"
    sheet.freeze_panes = "A4"
    for merged_range in list(sheet.merged_cells.ranges):
        sheet.unmerge_cells(str(merged_range))
    sheet.merge_cells(f"A1:{visible_end_column}1")
    sheet.merge_cells(f"A2:{visible_end_column}2")

    sheet["A1"] = WORKBOOK_TITLE
    sheet["A1"].fill = title_fill
    sheet["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")

    sheet["A2"] = get_week_label(active_week_start)
    sheet["A2"].font = Font(bold=True)
    sheet["A2"].alignment = Alignment(horizontal="center", vertical="center")

    visible_headers = ["Name", "TID", *get_weekday_headers(active_week_start)]
    for column, header in enumerate(visible_headers + METADATA_HEADERS, start=1):
        cell = sheet.cell(row=3, column=column, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        cell.fill = header_fill if column <= len(VISIBLE_HEADERS) else metadata_fill

    for column in range(1, len(SHEET_HEADERS) + 1):
        column_letter = get_column_letter(column)
        sheet.column_dimensions[column_letter].hidden = column > len(VISIBLE_HEADERS)

        if column == NAME_COLUMN:
            sheet.column_dimensions[column_letter].width = 24
        elif column == TID_COLUMN:
            sheet.column_dimensions[column_letter].width = 12
        elif column <= len(VISIBLE_HEADERS):
            sheet.column_dimensions[column_letter].width = 14
        else:
            sheet.column_dimensions[column_letter].width = 18


def style_data_row(sheet, row_number):
    thin_side = Side(style="thin", color="808080")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for column in range(1, len(SHEET_HEADERS) + 1):
        cell = sheet.cell(row=row_number, column=column)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sheet.cell(row=row_number, column=NAME_COLUMN).alignment = Alignment(horizontal="left", vertical="center")


def is_plan_format(sheet):
    first_two_headers = [sheet.cell(row=3, column=column).value for column in range(1, 3)]
    return sheet["A1"].value == WORKBOOK_TITLE and first_two_headers == ["Name", "TID"]


def create_workbook(week_start=None):
    workbook = Workbook()
    sheet = workbook.active
    active_week_start = week_start or next_week_range()[0]
    active_week_end = active_week_start + timedelta(days=6)
    style_worksheet(sheet, active_week_start)
    write_roster_rows(sheet, active_week_start, active_week_end)
    workbook.save(get_range_file(active_week_start, active_week_end))
    return workbook


def get_workbook():
    data_file = get_latest_data_file()
    if not data_file:
        return None

    workbook = load_workbook(data_file)
    if is_plan_format(workbook.active):
        week_start = get_sheet_week_start(workbook.active) or next_week_range()[0]
        style_worksheet(workbook.active, week_start)
        return workbook

    return None


def get_sheet_week_start(sheet):
    for row_number in range(4, sheet.max_row + 1):
        week_start = sheet.cell(row=row_number, column=WEEK_START_COLUMN).value
        if week_start:
            return datetime.strptime(str(week_start), "%Y-%m-%d").date()
    return None


def get_sheet_week_end(sheet):
    for row_number in range(4, sheet.max_row + 1):
        week_end = sheet.cell(row=row_number, column=WEEK_END_COLUMN).value
        if week_end:
            return datetime.strptime(str(week_end), "%Y-%m-%d").date()
    return None


def write_roster_rows(sheet, week_start, week_end, selections=None):
    selections = selections or set()
    employees = get_employees()

    if sheet.max_row > 3:
        sheet.delete_rows(4, sheet.max_row - 3)

    for employee_index, employee in enumerate(employees):
        row_number = sheet.max_row + 1
        sheet.cell(row=row_number, column=NAME_COLUMN, value=employee["name"])
        sheet.cell(row=row_number, column=TID_COLUMN, value=employee["tid"])

        for weekday_index, column in WEEKDAY_COLUMNS.items():
            value = "WFH" if (employee_index, weekday_index) in selections else ""
            sheet.cell(row=row_number, column=column, value=value)

        sheet.cell(row=row_number, column=METADATA_START_COLUMN, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        sheet.cell(row=row_number, column=EMAIL_COLUMN, value=employee.get("email", ""))
        sheet.cell(row=row_number, column=WEEK_START_COLUMN, value=week_start.isoformat())
        sheet.cell(row=row_number, column=WEEK_END_COLUMN, value=week_end.isoformat())
        sheet.cell(row=row_number, column=STATUS_COLUMN, value="Approved")
        style_data_row(sheet, row_number)


def parse_selection(selection):
    employees = get_employees()

    try:
        employee_index_text, weekday_index_text = selection.split("|", 1)
        employee_index = int(employee_index_text)
        weekday_index = int(weekday_index_text)
    except ValueError:
        return None

    if employee_index < 0 or employee_index >= len(employees):
        return None
    if weekday_index not in WEEKDAY_COLUMNS:
        return None
    return employee_index, weekday_index


def submit_weekly_wfh_plan(selected_values, week_start_text, week_end_text):
    week_start, week_end = parse_week_range(week_start_text, week_end_text)
    date_error = validate_week_range(week_start, week_end)
    if date_error:
        return RequestResult(date_error, "error")

    selections = set()
    daily_counts = {weekday_index: 0 for weekday_index in WEEKDAY_COLUMNS}

    for selected_value in selected_values:
        selection = parse_selection(selected_value)
        if selection is None:
            return RequestResult("Invalid table selection found. Please refresh and try again.", "error")

        if selection in selections:
            continue

        selections.add(selection)
        daily_counts[selection[1]] += 1

    over_limit_days = [
        get_weekday_headers(week_start)[weekday_index]
        for weekday_index, count in daily_counts.items()
        if count > DAILY_WFH_LIMIT
    ]
    if over_limit_days:
        return RequestResult(
            f"Only {DAILY_WFH_LIMIT} members can select WFH per day. Please reduce: {', '.join(over_limit_days)}.",
            "error",
        )

    workbook = Workbook()
    sheet = workbook.active
    style_worksheet(sheet, week_start)
    write_roster_rows(sheet, week_start, week_end, selections)
    workbook.save(get_range_file(week_start, week_end))
    return RequestResult(f"Weekly work from home plan has been saved to {get_range_file(week_start, week_end).name}.")


def get_selected_cells(sheet=None):
    if sheet is None:
        workbook = get_workbook()
        if workbook is None:
            return set()
        sheet = workbook.active

    selected = set()

    for employee_index, row_number in enumerate(range(4, sheet.max_row + 1)):
        for weekday_index, column in WEEKDAY_COLUMNS.items():
            if sheet.cell(row=row_number, column=column).value == "WFH":
                selected.add((employee_index, weekday_index))

    return selected


def get_wfh_selection_table(week_start_text=None, week_end_text=None):
    employees = get_employees()
    selected = set()
    requested_week_start = parse_date(week_start_text)
    requested_week_end = parse_date(week_end_text)

    if requested_week_start and not requested_week_end:
        requested_week_end = requested_week_start + timedelta(days=6)
    elif requested_week_end and not requested_week_start:
        requested_week_start = requested_week_end - timedelta(days=6)

    week_start = requested_week_start
    week_end = requested_week_end
    dates = get_working_week_dates(week_start) if week_start else []
    daily_counts = {
        weekday_index: sum(1 for selection in selected if selection[1] == weekday_index)
        for weekday_index in WEEKDAY_COLUMNS
    }

    rows = []
    for employee_index, employee in enumerate(employees):
        rows.append({
            "index": employee_index,
            "name": employee["name"],
            "tid": employee["tid"],
            "days": [
                {
                    "index": weekday_index,
                    "checked": (employee_index, weekday_index) in selected,
                }
                for weekday_index in WEEKDAY_COLUMNS
            ],
        })

    return {
        "title": WORKBOOK_TITLE,
        "week": get_week_label(week_start) if week_start else "",
        "week_start": week_start.isoformat() if week_start else "",
        "week_end": week_end.isoformat() if week_end else "",
        "dates": [
            {
                "label": day.strftime("%a") if week_start else WEEKDAY_LABELS[index],
                "date": day.strftime("%d-%b") if week_start else "",
                "count": daily_counts[index],
            }
            for index, day in enumerate(dates or [None] * len(WEEKDAY_LABELS))
        ],
        "rows": rows,
        "daily_limit": DAILY_WFH_LIMIT,
    }


def append_user_to_active_plan(employee):
    workbook = get_workbook()
    if workbook is None:
        return

    sheet = workbook.active
    week_start = get_sheet_week_start(sheet) or next_week_range()[0]
    week_end = get_sheet_week_end(sheet) or week_start + timedelta(days=6)
    row_number = sheet.max_row + 1

    sheet.cell(row=row_number, column=NAME_COLUMN, value=employee["name"])
    sheet.cell(row=row_number, column=TID_COLUMN, value=employee["tid"])
    sheet.cell(row=row_number, column=METADATA_START_COLUMN, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    sheet.cell(row=row_number, column=EMAIL_COLUMN, value=employee.get("email", ""))
    sheet.cell(row=row_number, column=WEEK_START_COLUMN, value=week_start.isoformat())
    sheet.cell(row=row_number, column=WEEK_END_COLUMN, value=week_end.isoformat())
    sheet.cell(row=row_number, column=STATUS_COLUMN, value="Approved")
    style_data_row(sheet, row_number)
    workbook.save(get_range_file(week_start, week_end))


def update_active_plan_user(original_employee, updated_employee):
    workbook = get_workbook()
    if workbook is None:
        return

    sheet = workbook.active
    week_start = get_sheet_week_start(sheet) or next_week_range()[0]
    week_end = get_sheet_week_end(sheet) or week_start + timedelta(days=6)

    for row_number in range(4, sheet.max_row + 1):
        name = normalize_text(sheet.cell(row=row_number, column=NAME_COLUMN).value).upper()
        tid = normalize_text(sheet.cell(row=row_number, column=TID_COLUMN).value).upper()

        if name == original_employee["name"].upper() and tid == original_employee["tid"].upper():
            sheet.cell(row=row_number, column=NAME_COLUMN, value=updated_employee["name"])
            sheet.cell(row=row_number, column=TID_COLUMN, value=updated_employee["tid"])
            sheet.cell(row=row_number, column=EMAIL_COLUMN, value=updated_employee.get("email", ""))
            style_data_row(sheet, row_number)
            workbook.save(get_range_file(week_start, week_end))
            return


def delete_active_plan_user(employee):
    workbook = get_workbook()
    if workbook is None:
        return

    sheet = workbook.active
    week_start = get_sheet_week_start(sheet) or next_week_range()[0]
    week_end = get_sheet_week_end(sheet) or week_start + timedelta(days=6)

    for row_number in range(sheet.max_row, 3, -1):
        name = normalize_text(sheet.cell(row=row_number, column=NAME_COLUMN).value).upper()
        tid = normalize_text(sheet.cell(row=row_number, column=TID_COLUMN).value).upper()

        if name == employee["name"].upper() and tid == employee["tid"].upper():
            sheet.delete_rows(row_number, 1)
            workbook.save(get_range_file(week_start, week_end))
            return


def add_user(name, tid="", email=""):
    name = normalize_text(name).upper()
    tid = normalize_text(tid).upper()
    email = normalize_text(email)

    if not name:
        return RequestResult("Name is required.", "error"), None

    employees = get_employees()
    duplicate_name = any(employee["name"].upper() == name for employee in employees)
    duplicate_tid = tid and any(employee["tid"].upper() == tid for employee in employees)

    if duplicate_name:
        return RequestResult("A user with this name already exists.", "error"), None
    if duplicate_tid:
        return RequestResult("A user with this TID already exists.", "error"), None

    workbook = get_users_workbook()
    sheet = workbook.active
    employee = {"name": name, "tid": tid, "email": email}
    sheet.append([employee["name"], employee["tid"], employee["email"]])
    workbook.save(USERS_FILE)
    append_user_to_active_plan(employee)

    return RequestResult(f"{name} has been added to the user roster."), employee


def update_user(user_index, name, tid="", email=""):
    name = normalize_text(name).upper()
    tid = normalize_text(tid).upper()
    email = normalize_text(email)

    if not name:
        return RequestResult("Name is required.", "error"), None

    workbook = get_users_workbook()
    sheet = workbook.active
    row_number = get_user_sheet_row(sheet, user_index)

    if row_number is None:
        return RequestResult("User was not found.", "error"), None

    employees = get_employees()
    original_employee = employees[user_index]
    duplicate_name = any(
        index != user_index and employee["name"].upper() == name
        for index, employee in enumerate(employees)
    )
    duplicate_tid = tid and any(
        index != user_index and employee["tid"].upper() == tid
        for index, employee in enumerate(employees)
    )

    if duplicate_name:
        return RequestResult("A user with this name already exists.", "error"), None
    if duplicate_tid:
        return RequestResult("A user with this TID already exists.", "error"), None

    updated_employee = {"name": name, "tid": tid, "email": email}
    sheet.cell(row=row_number, column=1, value=updated_employee["name"])
    sheet.cell(row=row_number, column=2, value=updated_employee["tid"])
    sheet.cell(row=row_number, column=3, value=updated_employee["email"])
    workbook.save(USERS_FILE)
    update_active_plan_user(original_employee, updated_employee)

    return RequestResult(f"{name} has been updated."), updated_employee


def delete_user(user_index):
    workbook = get_users_workbook()
    sheet = workbook.active
    row_number = get_user_sheet_row(sheet, user_index)

    if row_number is None:
        return RequestResult("User was not found.", "error")

    employees = get_employees()
    employee = employees[user_index]
    sheet.delete_rows(row_number, 1)
    workbook.save(USERS_FILE)
    delete_active_plan_user(employee)

    return RequestResult(f"{employee['name']} has been deleted.")


def get_wfh_plan_table():
    workbook = get_workbook()
    if workbook is None:
        rows = [
            [employee["name"], employee["tid"], *["" for _ in WEEKDAY_LABELS]]
            for employee in get_employees()
        ]
        return {
            "title": WORKBOOK_TITLE,
            "week": "",
            "headers": ["Name", "TID", *WEEKDAY_LABELS],
            "rows": rows,
        }

    sheet = workbook.active
    week_start = get_sheet_week_start(sheet) or next_week_range()[0]
    headers = ["Name", "TID", *get_weekday_headers(week_start)]
    rows = []

    for row in sheet.iter_rows(min_row=4, max_col=len(VISIBLE_HEADERS), values_only=True):
        row_values = [value or "" for value in row]
        if any(row_values):
            rows.append(row_values)

    return {
        "title": sheet["A1"].value or WORKBOOK_TITLE,
        "week": sheet["A2"].value or get_week_label(week_start),
        "headers": headers,
        "rows": rows,
    }


def fill_empty_weekdays_with_wfo():
    workbook = get_workbook()
    if workbook is None:
        return

    sheet = workbook.active
    week_start = get_sheet_week_start(sheet) or next_week_range()[0]
    week_end = get_sheet_week_end(sheet) or week_start + timedelta(days=6)

    for row_number in range(4, sheet.max_row + 1):
        name = sheet.cell(row=row_number, column=NAME_COLUMN).value
        if not name:
            continue

        for column in WEEKDAY_COLUMNS.values():
            cell = sheet.cell(row=row_number, column=column)
            if not cell.value:
                cell.value = "WFO"

        style_data_row(sheet, row_number)

    workbook.save(get_range_file(week_start, week_end))


def delete_wfh_excel_file():
    data_file = get_latest_data_file()
    if data_file and data_file.exists():
        data_file.unlink()


def get_wfh_plan_emails():
    workbook = get_workbook()
    if workbook is None:
        return []

    sheet = workbook.active
    emails = []

    for row_number in range(4, sheet.max_row + 1):
        email = sheet.cell(row=row_number, column=EMAIL_COLUMN).value or ""
        if "@" in email:
            emails.append(email)

    return sorted(set(emails))
